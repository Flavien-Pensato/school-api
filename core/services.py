"""Business logic spanning multiple models: student import, rotation."""

import csv
import io
import re
import unicodedata
from collections import Counter
from datetime import date

from django.db import transaction
from django.db.models import Prefetch, Q

from .models import Assignment, ClassPresence, Enrollment, Group, Student, Task

# Accepted column headers, compared after normalisation (lowercase, accents
# stripped, punctuation reduced to single spaces). School exports label the
# student columns "Elève Nom" / "Elève Prénom".
HEADER_ALIASES = {
    'first_name': {
        'first name', 'firstname', 'prenom', 'eleve prenom', 'prenom eleve',
        'prenom de l eleve',
    },
    'last_name': {
        'last name', 'lastname', 'nom', 'eleve nom', 'nom eleve',
        'nom de famille', 'nom de l eleve',
    },
    'external_id': {'external id', 'externalid', 'identifiant', 'id', 'ine'},
}
REQUIRED_FIELDS = ('first_name', 'last_name')
# Exports often carry a title banner ("MFR CHATTE", "ANNEE 2026/2027") above
# the real header row, so the header is searched for instead of assumed first.
HEADER_SEARCH_ROWS = 20


class ImportError_(Exception):
    """Import failed; `errors` is a list of {"row": n, "errors": [...]}
    or {"errors": [...]} for file-level problems."""

    def __init__(self, errors):
        self.errors = errors
        super().__init__(str(errors))


def _normalize(raw):
    """Lowercase, strip accents, collapse punctuation to single spaces."""
    text = unicodedata.normalize('NFKD', '' if raw is None else str(raw))
    text = ''.join(char for char in text if not unicodedata.combining(char))
    return ' '.join(part for part in re.split(r'[^a-z0-9]+', text.lower()) if part)


def _map_headers(raw_headers):
    """Map one row's cells to model fields. Returns {field: column_index}."""
    mapping = {}
    for index, raw in enumerate(raw_headers):
        name = _normalize(raw)
        if not name:
            continue
        for field, aliases in HEADER_ALIASES.items():
            if name in aliases and field not in mapping:
                mapping[field] = index
    return mapping


def _find_header_row(rows):
    """Find the header among the first rows. Returns (position, mapping)."""
    for position, (_, row) in enumerate(rows[:HEADER_SEARCH_ROWS]):
        mapping = _map_headers(row)
        if all(field in mapping for field in REQUIRED_FIELDS):
            return position, mapping
    raise ImportError_([{
        'errors': [
            'No header row found in the first '
            f'{HEADER_SEARCH_ROWS} non-empty rows. Required column(s): '
            f'{", ".join(REQUIRED_FIELDS)}. Accepted headers: '
            + '; '.join(
                f'{field}: {sorted(aliases)}'
                for field, aliases in HEADER_ALIASES.items()
            )
        ]
    }])


def _rows_from_csv(file):
    text = io.TextIOWrapper(file, encoding='utf-8-sig')
    sample = text.read(4096)
    text.seek(0)
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t')
    except csv.Error:
        dialect = csv.excel  # comma fallback
    return list(csv.reader(text, dialect))


def _rows_from_xlsx(file):
    from openpyxl import load_workbook

    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    return [
        ['' if cell is None else str(cell) for cell in row]
        for row in sheet.iter_rows(values_only=True)
    ]


def parse_student_file(file, filename):
    """Returns list of {'first_name', 'last_name', 'external_id', 'row'} dicts.
    Raises ImportError_ with row-level errors — nothing is written."""
    lowered = filename.lower()
    if lowered.endswith('.csv'):
        raw_rows = _rows_from_csv(file)
    elif lowered.endswith('.xlsx'):
        raw_rows = _rows_from_xlsx(file)
    else:
        raise ImportError_([{'errors': ['Unsupported file type; use .csv or .xlsx.']}])

    # Keep the real file line number so reported errors point at the right row.
    rows = [
        (line_number, row)
        for line_number, row in enumerate(raw_rows, start=1)
        if any((cell or '').strip() for cell in row)
    ]
    if not rows:
        raise ImportError_([{'errors': ['File is empty.']}])

    header_position, mapping = _find_header_row(rows)
    students, errors = [], []
    for line_number, row in rows[header_position + 1:]:
        def cell(field):
            index = mapping.get(field)
            if index is None or index >= len(row):
                return ''
            return (row[index] or '').strip()

        record = {field: cell(field) for field in HEADER_ALIASES}
        record['row'] = line_number
        row_errors = [
            f'{field} missing' for field in REQUIRED_FIELDS if not record[field]
        ]
        if row_errors:
            errors.append({'row': line_number, 'errors': row_errors})
        else:
            students.append(record)
    if errors:
        raise ImportError_(errors)
    return students


@transaction.atomic
def import_students(school_class, records):
    """Create/reuse students and enroll them in `school_class`.

    Matching rule: by external_id when present, else by case-insensitive
    (school, first_name, last_name). A student already enrolled in a
    DIFFERENT class this year is a row error — moving requires an explicit
    PATCH on the enrollment, never a silent reassign. All-or-nothing.
    """
    school = school_class.school_year.school
    year = school_class.school_year
    counts = {
        'created_students': 0,
        'reused_students': 0,
        'enrollments_created': 0,
        'already_enrolled': 0,
    }
    errors = []

    for index, record in enumerate(records, start=2):
        line_number = record.get('row', index)
        student = None
        if record['external_id']:
            student = Student.objects.filter(
                school=school, external_id=record['external_id']
            ).first()
        if student is None:
            student = Student.objects.filter(
                school=school,
                first_name__iexact=record['first_name'],
                last_name__iexact=record['last_name'],
            ).first()

        if student is None:
            student = Student.objects.create(
                school=school,
                first_name=record['first_name'],
                last_name=record['last_name'],
                external_id=record['external_id'],
            )
            counts['created_students'] += 1
        else:
            counts['reused_students'] += 1

        enrollment = student.enrollments.filter(school_year=year).first()
        if enrollment is None:
            Enrollment.objects.create(
                student=student, school_year=year, school_class=school_class
            )
            counts['enrollments_created'] += 1
        elif enrollment.school_class_id == school_class.pk:
            counts['already_enrolled'] += 1
        else:
            errors.append({
                'row': line_number,
                'errors': [
                    f'{student} already enrolled in '
                    f'{enrollment.school_class.name} this year; '
                    'move via the enrollments API.'
                ],
            })

    if errors:
        raise ImportError_(errors)  # rolls back the transaction
    return counts


def _min_cost_matching(groups, tasks, pair):
    """Assign each task to a distinct group minimizing the total of
    pair[(group.pk, task.pk)] — i.e. as few repeated (group, task) pairs
    as possible. Exact DP over a bitmask of groups; sizes are small because
    the working pool is already trimmed to len(tasks).

    Deterministic: cost ties are broken by task order (pk) then lowest
    group pk. If there are fewer groups than tasks, the leftover tasks stay
    unassigned. Returns [(group, task)].
    """
    skips_allowed = max(0, len(tasks) - len(groups))
    memo = {}

    def solve(task_index, used_mask, skips):
        if task_index == len(tasks):
            return (0, (), ())
        key = (task_index, used_mask, skips)
        if key in memo:
            return memo[key]
        task = tasks[task_index]
        best = None
        for bit, group in enumerate(groups):
            if used_mask & (1 << bit):
                continue
            cost, pks, picks = solve(task_index + 1, used_mask | (1 << bit), skips)
            candidate = (
                cost + pair[(group.pk, task.pk)],
                (group.pk,) + pks,
                ((group, task),) + picks,
            )
            if best is None or candidate[:2] < best[:2]:
                best = candidate
        if skips < skips_allowed:
            cost, pks, picks = solve(task_index + 1, used_mask, skips + 1)
            candidate = (cost, (float('inf'),) + pks, picks)
            if best is None or candidate[:2] < best[:2]:
                best = candidate
        memo[key] = best
        return best

    return list(solve(0, 0, 0)[2])


def _assign_class_cleaning(week, groups, manual_task_ids, total, pair, last_assigned):
    """Phase 0 — every class on site cleans its own room.

    The chore is implicit: it exists because the class is present, nobody
    enters it. It is carried by one group that has students in that class —
    a group is on site with its own class, and cleaning someone else's room
    is not the deal.

    Classes are served most-constrained first (fewest candidate groups), so a
    small class is not left empty-handed by a big one taking its only group.
    Within a class the pick follows the same fairness shape as the rotation:
    the group that did this room least often, then the one that worked least
    overall, then rested longest, then lowest pk.

    The picked groups are removed from the pool: `Assignment` allows one task
    per group per week, and a group scrubbing its classroom has done its
    share.

    Returns (picks, explanation, remaining_groups).
    """
    tasks_by_class = {
        task.school_class_id: task
        for task in Task.objects.filter(
            school_class__presences__week=week, is_active=True
        ).select_related('school_class')
        if task.pk not in manual_task_ids
    }
    # The class is present, so every enrollment in it is on site — no need to
    # re-check presence per student.
    groups_by_class = {}
    for class_id, group_id in Enrollment.objects.filter(
        school_class__presences__week=week, group__isnull=False
    ).values_list('school_class_id', 'group_id').distinct():
        groups_by_class.setdefault(class_id, set()).add(group_id)

    available = {group.pk: group for group in groups}
    picks, explanation = [], []
    ordered = sorted(
        tasks_by_class.items(),
        key=lambda item: (
            len(groups_by_class.get(item[0], ())),
            item[1].school_class.position or 0,
            item[1].school_class.name,
        ),
    )
    for class_id, task in ordered:
        candidates = [
            available[pk]
            for pk in groups_by_class.get(class_id, ())
            if pk in available
        ]
        if not candidates:
            explanation.append(
                f'{task.name}: unassigned (no group of that class is free)'
            )
            continue
        group = min(
            candidates,
            key=lambda g: (
                pair[(g.pk, task.pk)],
                total[g.pk],
                last_assigned.get(g.pk, date.min),
                g.pk,
            ),
        )
        del available[group.pk]
        picks.append((group, task))
        explanation.append(
            f'{group.name} → {task.name}: its class is on site, done '
            f'{pair[(group.pk, task.pk)]}× before'
        )
    return picks, explanation, [g for g in groups if g.pk in available]


@transaction.atomic
def generate_week_assignments(week):
    """Auto-assign each active task to one eligible group for `week`.

    Eligible groups: groups with at least one member whose class has a
    ClassPresence for this week.
    Manual assignments (is_manual=True) are preserved; their task and group
    are removed from the pools. Previous auto assignments for this week are
    replaced — re-running is idempotent.

    Fairness, over prior weeks of the same school year:
    - Phase 0 (class cleaning): every class on site has its room cleaned by
      one of its own groups; that group then rests from the rotation.
    - Phase 1 (rest fairness): if more groups than tasks, the groups with
      the fewest total assignments work first (tie: rested longest, then pk).
    - Phase 2 (pair fairness): greedy min-cost matching on how often each
      (group, task) pair already happened (tie: task pk, then group pk).

    Returns {"assignments": [Assignment], "explanation": [str]}.
    """
    school = week.school_year.school

    manual = list(
        week.assignments.filter(is_manual=True).select_related('task', 'group')
    )
    manual_task_ids = {a.task_id for a in manual}
    manual_group_ids = {a.group_id for a in manual}

    tasks = [
        t for t in Task.objects.filter(
            school=school, is_active=True, school_class__isnull=True
        )
        .order_by('pk')
        if t.pk not in manual_task_ids
    ]
    groups = [
        g for g in Group.objects.filter(
            school_year=week.school_year,
            enrollments__school_class__presences__week=week,
        )
        .distinct()
        .order_by('pk')
        if g.pk not in manual_group_ids
    ]

    # History: prior weeks of the same school year (manual rows count too —
    # a group that worked, worked).
    history = Assignment.objects.filter(
        week__school_year=week.school_year,
        week__start_date__lt=week.start_date,
    ).values_list('group_id', 'task_id', 'week__start_date')
    total = Counter()
    pair = Counter()
    last_assigned = {}
    for group_id, task_id, week_start in history:
        total[group_id] += 1
        pair[(group_id, task_id)] += 1
        if group_id not in last_assigned or week_start > last_assigned[group_id]:
            last_assigned[group_id] = week_start

    explanation = [
        f'{a.group.name} → {a.task.name}: manual assignment (kept)'
        for a in manual
    ]

    # Phase 0 — class cleaning, before anything else: it is the only chore
    # with a fixed set of eligible groups, so it picks first and the rotation
    # works with what is left.
    class_picks, class_explanation, groups = _assign_class_cleaning(
        week, groups, manual_task_ids, total, pair, last_assigned
    )
    explanation.extend(class_explanation)

    # Phase 1 — who works this week (rest fairness).
    working = sorted(
        groups,
        key=lambda g: (total[g.pk], last_assigned.get(g.pk, date.min), g.pk),
    )[: len(tasks)]
    for group in groups:
        if group not in working:
            explanation.append(
                f'{group.name} rests this week '
                f'({total[group.pk]} assignments so far — highest)'
            )

    # Phase 2 — who does what (pair fairness): exact min-cost matching.
    # Greedy is not enough — the last free slot can force a repeated pair
    # while a different arrangement avoids all repeats.
    picks = _min_cost_matching(working, tasks, pair)
    assigned_task_ids = set()
    for group, task in picks:
        assigned_task_ids.add(task.pk)
        explanation.append(
            f'{group.name} → {task.name}: done {pair[(group.pk, task.pk)]}× '
            'before (minimizes repeats this week)'
        )
    for task in tasks:
        if task.pk not in assigned_task_ids:
            explanation.append(f'{task.name}: unassigned (no group available)')

    week.assignments.filter(is_manual=False).delete()
    created = Assignment.objects.bulk_create(
        Assignment(week=week, task=task, group=group, is_manual=False)
        for group, task in class_picks + picks
    )
    return {'assignments': manual + created, 'explanation': explanation}


def build_week_dashboard(week):
    """Flat view of a week: every group on duty, with its task and its members.
    Shared by the JSON dashboard endpoint and the printable PDF.

    A group mixes classes, so only the members whose class is present are
    listed — the others are not on site to do the chore. Class cleaning rows
    lead; the rest keep the model's numeric ordering (1, 2, … 10).
    """
    assignments = {
        a.group_id: a
        for a in week.assignments.select_related('task', 'task__school_class')
    }
    present_members = Enrollment.objects.filter(
        school_class__presences__week=week
    ).select_related('student', 'school_class')
    groups = (
        Group.objects.filter(
            school_year=week.school_year,
            enrollments__school_class__presences__week=week,
        )
        .distinct()
        .prefetch_related(
            Prefetch(
                'enrollments', queryset=present_members, to_attr='on_site'
            )
        )
    )
    rows = []
    for group in groups:
        assignment = assignments.get(group.pk)
        rows.append({
            '_is_class_task': bool(
                assignment and assignment.task.school_class_id
            ),
            'id': group.pk,
            'name': group.name,
            'students': [
                {
                    'id': e.student.pk,
                    'first_name': e.student.first_name,
                    'last_name': e.student.last_name,
                    'school_class': e.school_class.name,
                }
                for e in sorted(
                    group.on_site,
                    key=lambda e: (e.student.last_name, e.student.first_name),
                )
            ],
            'task': (
                {
                    'id': assignment.task.pk,
                    'name': assignment.task.name,
                    # Set only for a class's own cleaning: lets the sheet
                    # group those rows apart from the rotating chores.
                    'school_class': (
                        assignment.task.school_class.name
                        if assignment.task.school_class_id else None
                    ),
                }
                if assignment else None
            ),
        })
    # Cleaning your own room comes first on the sheet: it is the line a
    # class looks for, and the one that moves every week with who is on site.
    rows.sort(key=lambda row: not row.pop('_is_class_task'))
    return {
        'week': {
            'id': week.pk,
            'start_date': week.start_date.isoformat(),
            'label': week.label,
        },
        'school': {'id': week.school.pk, 'name': week.school.name},
        'groups': rows,
    }


def build_year_stats(school_year):
    """Per-group fairness matrix for a school year: how often each group
    did each task, plus totals and rest counts."""
    # The school's rotating chores, plus the cleaning tasks of THIS year's
    # classes. Another year's class tasks belong to the same school but mean
    # nothing here — they would add a permanently empty column per class.
    tasks = sorted(
        Task.objects.filter(school=school_year.school)
        .filter(
            Q(school_class__isnull=True)
            | Q(school_class__school_year=school_year)
        )
        .select_related('school_class'),
        key=lambda task: (
            task.school_class_id is None,  # class cleaning columns first
            task.school_class.position or 0 if task.school_class_id else 0,
            task.pk,
        ),
    )
    task_names = {task.pk: task.name for task in tasks}
    # A group is on duty as soon as one member's class is present, and two
    # members can bring the same week in — count distinct (group, week) pairs.
    weeks_present = Counter(
        group_id
        for group_id, _week_id in Enrollment.objects.filter(
            school_year=school_year,
            group__isnull=False,
            school_class__presences__isnull=False,
        )
        .values_list('group_id', 'school_class__presences__week_id')
        .distinct()
    )
    group_rows = {
        group.pk: {
            'group': {'id': group.pk, 'name': group.name},
            'totals': {},
            'total': 0,
            'weeks_present': weeks_present[group.pk],
            'weeks_rested': 0,
        }
        for group in Group.objects.filter(school_year=school_year)
    }
    for group_id, task_id in Assignment.objects.filter(
        week__school_year=school_year
    ).values_list('group_id', 'task_id'):
        row = group_rows.get(group_id)
        if row is None:
            continue
        name = task_names.get(task_id, str(task_id))
        row['totals'][name] = row['totals'].get(name, 0) + 1
        row['total'] += 1
    for row in group_rows.values():
        row['weeks_rested'] = row['weeks_present'] - row['total']
    return {
        'school_year': {'id': school_year.pk, 'name': school_year.name},
        # The columns to render, in order — the client cannot rebuild this
        # list from /api/tasks/, which hides class tasks and spans years.
        'tasks': [
            {
                'id': task.pk,
                'name': task.name,
                'school_class': (
                    task.school_class.name if task.school_class_id else None
                ),
            }
            for task in tasks
        ],
        'groups': list(group_rows.values()),
    }


def revoke_class_presence(presence):
    """Un-check one cell of the presence grid: the class is no longer on-site
    that week, so any group it emptied has no work that week either.

    Deleting the ClassPresence row alone is not enough. Assignment rows for
    the groups that just lost their last on-site member survive, and nothing
    else clears them: generate_week_assignments only deletes is_manual=False
    rows, and only for groups it still considers present. Such orphans stay
    invisible in build_week_dashboard while build_year_stats keeps counting
    them into `total`, pushing `weeks_rested` negative.

    A group that still holds a member from another present class keeps its
    assignment — it can do the chore short-handed. The one exception is this
    class's own cleaning task: the room is empty, so it is not cleaned, and
    that row goes whatever else the group still has on site.
    Manual assignments are not spared when the group does empty out: nobody on site cannot work, whoever
    typed the override. Re-checking the box leaves those tasks un-assigned;
    run the rotation again to refill them.
    """
    with transaction.atomic():
        week = presence.week
        school_class = presence.school_class
        presence.delete()
        # The room is not used this week, so it is not cleaned -- even though
        # the group that had the job usually still has members on site with
        # another class, which is exactly what the query below spares.
        cleaning, _ = Assignment.objects.filter(
            week=week, task__school_class=school_class
        ).delete()
        orphans, _ = Assignment.objects.filter(week=week).exclude(
            group__enrollments__school_class__presences__week=week
        ).delete()
    return cleaning + orphans


def build_year_presence_grid(school_year):
    """Presence table for a whole school year: one row per class, one column
    per week, for the UI to render checkboxes.

    Each cell is the ClassPresence pk (the class is present that week) or
    None (absent) — the pk is what the client needs to DELETE the row when
    the box is un-checked. `cells` is positional: index i is `weeks[i]`.
    """
    weeks = list(school_year.weeks.order_by('start_date'))
    column_of = {week.pk: index for index, week in enumerate(weeks)}
    presences = ClassPresence.objects.filter(
        week__school_year=school_year
    ).values_list('school_class_id', 'week_id', 'pk')
    cells_by_class = {}
    for class_id, week_id, presence_id in presences:
        row = cells_by_class.setdefault(class_id, [None] * len(weeks))
        row[column_of[week_id]] = presence_id
    return {
        'school_year': {'id': school_year.pk, 'name': school_year.name},
        'weeks': [
            {
                'id': week.pk,
                'start_date': week.start_date.isoformat(),
                'label': week.label,
            }
            for week in weeks
        ],
        'classes': [
            {
                'id': class_id,
                'name': name,
                'cells': cells_by_class.get(class_id, [None] * len(weeks)),
            }
            for class_id, name in school_year.classes.values_list(
                'pk', 'name'
            ).order_by('name')
        ],
    }
